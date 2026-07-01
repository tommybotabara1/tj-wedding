#!/usr/bin/env python3
"""
gws.py — Google Drive utility for TJ Wedding project.

Functions:
    download_workbook() → downloads the wedding xlsx from Drive, returns openpyxl.Workbook
    upload_workbook(wb) → uploads modified xlsx back to Drive (replaces existing)

Credentials:
    - credentials.json  : service account key (project root)
    - GOOGLE_DRIVE_FILE_ID in .env
"""

import io
import os
import sys

from dotenv import load_dotenv
from google.oauth2.service_account import Credentials
from googleapiclient.discovery import build
from googleapiclient.errors import HttpError
from googleapiclient.http import MediaIoBaseDownload, MediaIoBaseUpload

try:
    import openpyxl
except ImportError:
    print("ERROR: openpyxl not installed. Run: pip install openpyxl")
    sys.exit(1)

load_dotenv()

SCOPES     = [
    "https://www.googleapis.com/auth/drive",
    "https://www.googleapis.com/auth/spreadsheets.readonly",
]
CREDS_FILE = os.path.join(os.path.dirname(__file__), "..", "credentials.json")
FILE_ID    = os.environ.get("GOOGLE_DRIVE_FILE_ID", "")


def _creds():
    creds = Credentials.from_service_account_file(CREDS_FILE, scopes=SCOPES)
    return creds


def _drive_service():
    return build("drive", "v3", credentials=_creds())


def _sheets_service():
    return build("sheets", "v4", credentials=_creds())


def _append_values(ws, values):
    for row in values:
        ws.append(row)


def _download_workbook_from_sheets_api() -> openpyxl.Workbook:
    """Build a compact workbook from visible source tabs when Drive export is too large."""
    service = _sheets_service()
    ranges = [
        "'Timeline  Task List'!A1:D120",
        "'Budget + Vendor Tracker'!A1:I120",
        "'Guest List'!A1:K1000",
        "'Wedding Day Itinerary'!A1:G100",
    ]
    result = service.spreadsheets().values().batchGet(
        spreadsheetId=FILE_ID,
        ranges=ranges,
        valueRenderOption="FORMATTED_VALUE",
        dateTimeRenderOption="FORMATTED_STRING",
    ).execute()

    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    for item in result.get("valueRanges", []):
        range_name = item.get("range", "")
        sheet_name = range_name.split("!", 1)[0].strip("'")
        ws = wb.create_sheet(sheet_name)
        _append_values(ws, item.get("values", []))
    return wb


def download_workbook() -> openpyxl.Workbook:
    """Download the wedding xlsx from Drive and return an openpyxl Workbook."""
    if not FILE_ID:
        raise ValueError("GOOGLE_DRIVE_FILE_ID not set in .env")

    service = _drive_service()
    meta = service.files().get(fileId=FILE_ID, fields="mimeType").execute()
    if meta.get("mimeType") == "application/vnd.google-apps.spreadsheet":
        request = service.files().export_media(
            fileId=FILE_ID,
            mimeType="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
    else:
        request = service.files().get_media(fileId=FILE_ID)

    buf = io.BytesIO()
    downloader = MediaIoBaseDownload(buf, request)

    done = False
    try:
        while not done:
            _, done = downloader.next_chunk()
    except HttpError as exc:
        content = exc.content.decode("utf-8", errors="ignore") if getattr(exc, "content", None) else ""
        if "exportSizeLimitExceeded" in content:
            return _download_workbook_from_sheets_api()
        raise

    buf.seek(0)
    return openpyxl.load_workbook(buf, data_only=True)


def upload_workbook(wb: openpyxl.Workbook) -> str:
    """Upload a modified workbook back to Drive, replacing the existing file.

    Returns the file ID.
    """
    if not FILE_ID:
        raise ValueError("GOOGLE_DRIVE_FILE_ID not set in .env")

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    service = _drive_service()
    media = MediaIoBaseUpload(
        buf,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        resumable=False,
    )
    result = service.files().update(fileId=FILE_ID, media_body=media).execute()
    return result.get("id", FILE_ID)


def upload_workbook_bytes(path: str) -> str:
    """Upload a local .xlsx file's RAW bytes to Drive, replacing the existing file.

    Use this (not upload_workbook) when the file contains images/drawings: it
    streams the bytes verbatim, so nothing is lost. Requires the service account
    to have Editor access on FILE_ID.
    """
    if not FILE_ID:
        raise ValueError("GOOGLE_DRIVE_FILE_ID not set in .env")

    with open(path, "rb") as fh:
        buf = io.BytesIO(fh.read())
    buf.seek(0)

    service = _drive_service()
    media = MediaIoBaseUpload(
        buf,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        resumable=False,
    )
    result = service.files().update(fileId=FILE_ID, media_body=media).execute()
    return result.get("id", FILE_ID)


if __name__ == "__main__":
    print("Downloading workbook...")
    wb = download_workbook()
    print(f"Sheets: {wb.sheetnames}")
    print("Done.")
