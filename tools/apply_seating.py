#!/usr/bin/env python3
"""
apply_seating.py — Lossless write-back of the approved 10-table seating plan
into the Guest List sheet's "Table #" column (col K).

Same lossless principle as apply_savoy_updates.py: edit only the target cells
in the Guest List worksheet XML, preserve every other zip entry (66 images /
16 drawings) byte-for-byte. Existing Table# cells are replaced in place; the
~12 previously-blank cells are inserted into their <row> in column order.

Plan: 10 round tables of 12 (Savoy Connect Lounge round-table setup).
Guests are matched by a unique substring of their Name (col C).

Usage:
    python tools/apply_seating.py <input.xlsx> <output.xlsx>
"""

import re
import sys
import zipfile

import openpyxl
from openpyxl.utils import get_column_letter, column_index_from_string

from apply_savoy_updates import _sheet_path_map, _xml_escape

SHEET = "Guest List"

# table -> list of unique Name substrings
PLAN = {
    1:  ["Dean Joan", "Chairperson Christian", "Benedict Avila", "Karen Austria",
         "Shiela Chan", "Cheryl Imelda", "Patrick M. Santillan"],
    2:  ["Andrew Paul", "Analiza", "William G. Tan", "Jaime C. Perez",
         "Corazon", "Emalyn", "Yshmael", "Jennifer Calma"],
    3:  ["Joel M", "Olivia C", "Hannah C", "Joval C", "Romulo Fronda", "Ronulfo",
         "Mamerto", "Denise M", "JV Cruz", "Agustina", "Josiah", "Luis Montenegro"],
    4:  ["Elaine Bota", "Sherwin", "Teddy", "Sheryl De Dios", "Dong De Dios",
         "Jun Bota", "Jackie", "Mae Aspra", "Kervy", "Emil De Dios", "Luis De Dios"],
    5:  ["Alvin", "Bea Bota", "Nina", "Seansean", "Asuncion", "Dennis Mira",
         "Axl John", "George Suarez", "Ellwin", "John Goloya"],
    6:  ["Dimples", "Paulo Luis", "Donnabelle", "Josett", "Haidee", "Gemela",
         "Jaimini", "Ceijas", "Adrianne", "Julia Cecilia"],
    7:  ["Jan Carlo Ponce", "Louise", "Tricia", "Francis Valencia", "Didi Rico",
         "Cuenco", "Ed Paul", "Britanico", "Sharm", "Bryle", "Rahim", "Chito"],
    8:  ["Traviz", "Vaunn", "Jade Calera", "Michael Somosot", "Gabriel Ventura",
         "Patrick Soriano", "Betlee"],
    9:  ["Julian Christopher", "Michelle A", "Aldrine", "Micca", "Carl Tristan",
         "Vanessa", "Marielle", "Jean Distal", "Eloissa", "Joy Giron"],
    10: ["Abigail", "Paolo A. Macasaet", "Michaela Sta", "Godofredo",
         "John Carlo Soriaga", "Geraldine", "Jason Deichmann", "Piolo",
         "Kristian", "Chairperson Mark"],
}


def _set_cell(xml, ref, value, style=""):
    """Replace <c r=ref> if present (keep its style), else insert into the row."""
    pat = re.compile(r'<c r="' + re.escape(ref) + r'"([^>]*?)(?:/>|>.*?</c>)', re.DOTALL)
    m = pat.search(xml)
    if m:
        s = re.search(r'\bs="(\d+)"', m.group(1))
        s_attr = f' s="{s.group(1)}"' if s else (f' s="{style}"' if style else "")
        return xml[:m.start()] + f'<c r="{ref}"{s_attr}><v>{value}</v></c>' + xml[m.end():]
    # Insert into the row in column order.
    col_letters = re.match(r'[A-Z]+', ref).group(0)
    row_num = ref[len(col_letters):]
    target_idx = column_index_from_string(col_letters)
    s_attr = f' s="{style}"' if style else ""
    new_cell = f'<c r="{ref}"{s_attr}><v>{value}</v></c>'
    rm = re.search(r'(<row r="' + row_num + r'"[^>]*>)(.*?)(</row>)', xml, re.DOTALL)
    if not rm:
        raise ValueError(f"Row {row_num} not found to insert {ref}")
    body = rm.group(2)
    insert_at = len(body)
    for cm in re.finditer(r'<c r="([A-Z]+)\d+"', body):
        if column_index_from_string(cm.group(1)) > target_idx:
            insert_at = cm.start()
            break
    new_body = body[:insert_at] + new_cell + body[insert_at:]
    return xml[:rm.start()] + rm.group(1) + new_body + rm.group(3) + xml[rm.end():]


def build_edits(in_path):
    """Resolve each guest's row in col K and return [(ref, table)] + a style hint."""
    wb = openpyxl.load_workbook(in_path)
    ws = wb[SHEET]
    # header row + columns
    hdr = None
    for r in range(1, 8):
        vals = [str(ws.cell(r, c).value).strip().lower() if ws.cell(r, c).value else "" for c in range(1, 16)]
        if "name" in vals and "side" in vals:
            hdr = r
            cols = {vals[c - 1]: c for c in range(1, 16) if vals[c - 1]}
            break
    name_c = cols["name"]
    table_c = cols.get("table #") or cols.get("table#") or cols.get("table")
    tcol = get_column_letter(table_c)
    # style of an existing Table# cell (for inserted blanks)
    style = ""
    for r in range(hdr + 1, ws.max_row + 1):
        if ws.cell(r, table_c).value is not None:
            style = str(ws.cell(r, table_c)._style and ws.cell(r, table_c).style_id or "")
            break
    # name -> row
    rows = {}
    for r in range(hdr + 1, ws.max_row + 1):
        nm = ws.cell(r, name_c).value
        if nm:
            rows[str(nm).strip()] = r
    edits = []
    used = set()
    for table, subs in PLAN.items():
        for sub in subs:
            hits = [n for n in rows if sub.lower() in n.lower()]
            if len(hits) != 1:
                raise ValueError(f"Ambiguous/missing name '{sub}': {hits}")
            n = hits[0]
            if rows[n] in used:
                raise ValueError(f"Duplicate row for {n}")
            used.add(rows[n])
            edits.append((f"{tcol}{rows[n]}", table))
    return edits, str(style)


def apply_seating(in_path, out_path):
    edits, style = build_edits(in_path)
    with zipfile.ZipFile(in_path, "r") as zin:
        path = _sheet_path_map(zin)[SHEET]
        xml = zin.read(path).decode("utf-8")
        for ref, table in edits:
            xml = _set_cell(xml, ref, table, style)
        data = xml.encode("utf-8")
        with zipfile.ZipFile(out_path, "w", zipfile.ZIP_DEFLATED) as zout:
            for item in zin.infolist():
                zout.writestr(item, data if item.filename == path else zin.read(item.filename))
    print(f"Wrote {len(edits)} Table# assignments -> {out_path}")


if __name__ == "__main__":
    if len(sys.argv) != 3:
        print("Usage: python tools/apply_seating.py <input.xlsx> <output.xlsx>")
        sys.exit(1)
    apply_seating(sys.argv[1], sys.argv[2])
